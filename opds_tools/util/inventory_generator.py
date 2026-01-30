# opds_tools/util/inventory_generator.py

import csv
import io
import logging
import requests
from urllib.parse import urljoin
from xml.etree.ElementTree import Element, SubElement, tostring
from xml.dom import minidom
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

logger = logging.getLogger(__name__)


def crawl_feed_for_inventory(feed_url, max_pages=None, auth=None, username=None, password=None):
    """
    Crawl OPDS feed and extract inventory data.
    
    Args:
        feed_url: URL of the OPDS feed to crawl
        max_pages: Maximum number of pages to crawl (None for unlimited)
        auth: Tuple of (username, password) for basic auth
        username: Username for basic auth (alternative to auth tuple)
        password: Password for basic auth (alternative to auth tuple)
    
    Returns:
        dict: {
            'inventory': list of publication inventory records,
            'stats': summary statistics,
            'errors': list of error messages
        }
    """
    inventory = []
    errors = []
    pages_crawled = 0
    seen_urls = set()
    
    # Handle auth
    if username and password and not auth:
        auth = (username, password)
    
    def crawl_page(url):
        nonlocal pages_crawled
        
        if url in seen_urls:
            logger.warning(f"Already visited: {url}, skipping to avoid loop.")
            return
        
        if max_pages and pages_crawled >= max_pages:
            logger.info(f"Reached max_pages limit: {max_pages}")
            return
        
        seen_urls.add(url)
        
        try:
            logger.info(f"Crawling page {pages_crawled + 1}: {url}")
            response = requests.get(url, auth=auth, timeout=30)
            response.raise_for_status()
            feed_data = response.json()
            
            pages_crawled += 1
            
            # Extract publications from this page
            publications = feed_data.get('publications') or feed_data.get('items') or []
            logger.info(f"Found {len(publications)} publications on page {pages_crawled}")
            
            for pub in publications:
                try:
                    record = extract_inventory_record(pub, url)
                    if record:
                        inventory.append(record)
                except Exception as e:
                    error_msg = f"Error extracting publication: {str(e)}"
                    logger.warning(error_msg)
                    errors.append(error_msg)
            
            # Follow next link
            next_url = None
            for link in feed_data.get('links', []):
                if isinstance(link, dict) and link.get('rel') == 'next':
                    next_url = link.get('href')
                    if next_url:
                        next_url = urljoin(url, next_url)
                    break
            
            if next_url and (not max_pages or pages_crawled < max_pages):
                crawl_page(next_url)
                
        except requests.RequestException as e:
            error_msg = f"Network error on {url}: {str(e)}"
            logger.error(error_msg)
            errors.append(error_msg)
        except Exception as e:
            error_msg = f"Error processing {url}: {str(e)}"
            logger.error(error_msg)
            errors.append(error_msg)
    
    # Start crawling
    crawl_page(feed_url)
    
    # Calculate statistics
    stats = calculate_inventory_stats(inventory)
    stats['pages_crawled'] = pages_crawled
    stats['total_publications'] = len(inventory)
    stats['errors_count'] = len(errors)
    
    return {
        'inventory': inventory,
        'stats': stats,
        'errors': errors
    }


def extract_inventory_record(publication, base_url=None):
    """
    Extract inventory data from a single publication object.
    
    Args:
        publication: Publication object from OPDS feed
        base_url: Base URL for resolving relative links
    
    Returns:
        dict: Inventory record with identifier, title, author, publisher, format, drm
    """
    metadata = publication.get('metadata', {})
    
    # Extract identifier
    identifier = metadata.get('identifier', 'N/A')
    
    # Extract title
    title = metadata.get('title', 'Untitled')
    
    # Extract author
    author = extract_author(metadata)
    
    # Extract publisher
    publisher = extract_publisher(metadata)
    
    # Extract format and DRM from links
    links = publication.get('links', [])
    format_str = extract_format_from_links(links)
    drm_str = extract_drm_from_links(links)
    
    return {
        'identifier': identifier,
        'title': title,
        'author': author,
        'publisher': publisher,
        'format': format_str,
        'drm': drm_str
    }


def extract_author(metadata):
    """Extract author(s) from metadata."""
    raw_authors = metadata.get('authors') or metadata.get('author')
    
    if isinstance(raw_authors, str):
        return raw_authors
    elif isinstance(raw_authors, dict):
        return raw_authors.get('name', 'Unknown')
    elif isinstance(raw_authors, list):
        authors = []
        for a in raw_authors:
            if isinstance(a, dict):
                name = a.get('name', '')
                if name:
                    authors.append(name)
            elif isinstance(a, str):
                authors.append(a)
        return ', '.join(authors) if authors else 'Unknown'
    else:
        return 'Unknown'


def extract_publisher(metadata):
    """Extract publisher from metadata."""
    raw_publisher = metadata.get('publisher', '')
    
    if isinstance(raw_publisher, str):
        return raw_publisher if raw_publisher else 'Unknown'
    elif isinstance(raw_publisher, dict):
        return raw_publisher.get('name', 'Unknown')
    elif isinstance(raw_publisher, list):
        publishers = []
        for p in raw_publisher:
            if isinstance(p, dict):
                name = p.get('name', '')
                if name:
                    publishers.append(name)
            elif isinstance(p, str):
                publishers.append(p)
        return ', '.join(publishers) if publishers else 'Unknown'
    else:
        return 'Unknown'


def extract_format_from_links(links):
    """
    Determine format(s) from publication links.
    
    Args:
        links: List of link objects from publication
    
    Returns:
        str: Comma-separated format names (e.g., "EPUB", "EPUB, PDF")
    """
    formats = set()
    
    # MIME type to format mapping
    mime_to_format = {
        'application/epub+zip': 'EPUB',
        'application/pdf': 'PDF',
        'application/webpub+json': 'WebPub',
        'application/audiobook+json': 'Audiobook',
        'application/divina+json': 'DiViNa',
        'application/vnd.comicbook+zip': 'Comic',
        'application/x-mobipocket-ebook': 'MOBI',
    }
    
    for link in links:
        if not isinstance(link, dict):
            continue
        
        link_type = link.get('type', '')
        rel = link.get('rel', '')
        
        # Only consider acquisition links
        if not rel or 'acquisition' not in rel.lower():
            continue
        
        # Check for known MIME types
        for mime, format_name in mime_to_format.items():
            if mime in link_type:
                formats.add(format_name)
                break
    
    if formats:
        return ', '.join(sorted(formats))
    else:
        return 'Unknown'


def extract_drm_from_links(links):
    """
    Determine DRM type from publication links.
    
    Args:
        links: List of link objects from publication
    
    Returns:
        str: DRM type (e.g., "LCP DRM", "Adobe DRM", "Bearer Token", "None")
    """
    drm_types = set()
    has_acquisition = False
    
    for link in links:
        if not isinstance(link, dict):
            continue
        
        link_type = link.get('type', '').lower()
        rel = link.get('rel', '').lower()
        
        # Only consider acquisition links
        if not rel or 'acquisition' not in rel:
            continue
        
        has_acquisition = True
        
        # Check for LCP DRM
        if 'lcp' in link_type or 'readium.lcp' in link_type:
            drm_types.add('LCP DRM')
        
        # Check for Adobe DRM
        elif 'adobe' in link_type or 'adept' in link_type:
            drm_types.add('Adobe DRM')
        
        # Check for generic DRM indicators
        elif 'drm' in link_type or '+drm' in link_type:
            drm_types.add('DRM')
        
        # Check for Bearer Token in rel
        elif 'bearer-token' in rel or 'bearer_token' in rel:
            drm_types.add('Bearer Token')
    
    if drm_types:
        return ', '.join(sorted(drm_types))
    elif has_acquisition:
        return 'None'
    else:
        return 'Unknown'


def calculate_inventory_stats(inventory):
    """
    Calculate summary statistics from inventory data.
    
    Args:
        inventory: List of inventory records
    
    Returns:
        dict: Statistics including format and DRM breakdowns
    """
    format_counts = {}
    drm_counts = {}
    
    for record in inventory:
        # Count formats
        formats = record.get('format', 'Unknown')
        if formats in format_counts:
            format_counts[formats] += 1
        else:
            format_counts[formats] = 1
        
        # Count DRM types
        drm = record.get('drm', 'Unknown')
        if drm in drm_counts:
            drm_counts[drm] += 1
        else:
            drm_counts[drm] = 1
    
    # Sort by count
    format_stats = sorted(
        [{'type': k, 'count': v} for k, v in format_counts.items()],
        key=lambda x: x['count'],
        reverse=True
    )
    
    drm_stats = sorted(
        [{'type': k, 'count': v} for k, v in drm_counts.items()],
        key=lambda x: x['count'],
        reverse=True
    )
    
    return {
        'format_counts': format_stats,
        'drm_counts': drm_stats,
        'unique_formats': len(format_counts),
        'unique_drm_types': len(drm_counts)
    }


def generate_inventory_csv(inventory_data):
    """
    Generate CSV from inventory data.
    
    Args:
        inventory_data: List of inventory records
    
    Returns:
        str: CSV content
    """
    output = io.StringIO()
    fieldnames = ['identifier', 'title', 'author', 'publisher', 'format', 'drm']
    
    writer = csv.DictWriter(output, fieldnames=fieldnames, quoting=csv.QUOTE_MINIMAL)
    writer.writeheader()
    
    for record in inventory_data:
        # Ensure all fields exist
        row = {key: record.get(key, '') for key in fieldnames}
        writer.writerow(row)
    
    return output.getvalue()


def generate_inventory_xml(inventory_data):
    """
    Generate XML from inventory data.
    
    Args:
        inventory_data: List of inventory records
    
    Returns:
        str: Pretty-printed XML content
    """
    root = Element('inventory')
    root.set('total', str(len(inventory_data)))
    
    for record in inventory_data:
        pub_elem = SubElement(root, 'publication')
        
        # Add child elements for each field
        for field in ['identifier', 'title', 'author', 'publisher', 'format', 'drm']:
            child = SubElement(pub_elem, field)
            child.text = str(record.get(field, ''))
    
    # Pretty print
    xml_string = tostring(root, encoding='unicode')
    dom = minidom.parseString(xml_string)
    return dom.toprettyxml(indent="  ")


def generate_inventory_excel(inventory_data):
    """
    Generate Excel (.xlsx) file from inventory data.
    
    Args:
        inventory_data: List of inventory records
    
    Returns:
        bytes: Excel file content as bytes
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory"
    
    # Define header style
    header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Define cell alignment
    cell_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    
    # Column headers
    headers = ['identifier', 'title', 'author', 'publisher', 'format', 'drm']
    
    # Write headers with styling
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header.capitalize())
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Write data rows
    for row_idx, record in enumerate(inventory_data, start=2):
        for col_idx, field in enumerate(headers, start=1):
            value = record.get(field, '')
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = cell_alignment
    
    # Auto-adjust column widths
    ws.column_dimensions['A'].width = 25  # identifier
    ws.column_dimensions['B'].width = 40  # title
    ws.column_dimensions['C'].width = 25  # author
    ws.column_dimensions['D'].width = 25  # publisher
    ws.column_dimensions['E'].width = 25  # format
    ws.column_dimensions['F'].width = 25  # drm
    
    # Freeze the header row
    ws.freeze_panes = "A2"
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output.getvalue()
