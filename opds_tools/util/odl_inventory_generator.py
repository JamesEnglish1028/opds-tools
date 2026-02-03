# opds_tools/util/odl_inventory_generator.py

import csv
import io
import logging
import requests
from urllib.parse import urljoin
from xml.etree.ElementTree import Element, SubElement, tostring
from xml.dom import minidom
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from opds_tools.util.inventory_generator import extract_author, extract_publisher
from opds_tools.util.odl_analyzer import (
    detect_odl_formats,
    detect_odl_drm_scheme,
    extract_license_terms,
)

logger = logging.getLogger(__name__)


def crawl_odl_feed_for_inventory(feed_url, max_pages=None, auth=None, username=None, password=None, progress_callback=None):
    """
    Crawl ODL feed and extract inventory data.
    
    Supports two authentication methods:
    1. Basic Auth (HTTP Authorization header) - pass username & password
    2. Query Parameter Auth (for APIs like Feedbooks) - append ?key=value to URL

    Returns:
        dict: {
            'inventory': list of publication inventory records,
            'stats': summary statistics,
            'errors': list of error messages
        }
    """
    inventory = []
    errors = []
    pages_crawled = 0
    seen_urls = set()

    # Prepare auth
    auth_tuple = None
    if username and password and not auth:
        auth_tuple = (username, password)
        logger.info("Using basic auth credentials: %s", username)
    elif auth:
        auth_tuple = auth
        logger.info("Using provided auth tuple")

    current_auth = auth_tuple

    def crawl_page(url):
        nonlocal pages_crawled, current_auth

        if url in seen_urls:
            logger.warning("Already visited: %s, skipping to avoid loop.", url)
            return

        if max_pages and pages_crawled >= max_pages:
            logger.info("Reached max_pages limit: %s", max_pages)
            return

        seen_urls.add(url)

        try:
            logger.info("Crawling ODL page %s: %s", pages_crawled + 1, url)

            if progress_callback:
                progress_callback('page_started', {
                    'page_number': pages_crawled + 1,
                    'url': url,
                    'total_publications': len(inventory)
                })

            logger.debug("Making request to: %s with auth: %s", url, "enabled" if auth_tuple else "disabled")
            if auth_tuple:
                logger.debug("Auth tuple provided: (%s, [password])", auth_tuple[0])
            
            response = requests.get(url, auth=current_auth, timeout=30)
            logger.debug("Request headers sent: %s", dict(response.request.headers))
            response.raise_for_status()
            feed_data = response.json()
            logger.debug("Response status: %s, content length: %s", response.status_code, len(response.text))

            pages_crawled += 1

            publications = feed_data.get('publications') or feed_data.get('items') or []
            logger.info("Found %s publications on ODL page %s", len(publications), pages_crawled)

            if progress_callback:
                progress_callback('publications_found', {
                    'page_number': pages_crawled,
                    'count': len(publications),
                    'total_publications': len(inventory)
                })

            for pub in publications:
                try:
                    record = extract_odl_inventory_record(pub, url)
                    if record:
                        inventory.append(record)
                except Exception as e:
                    error_msg = f"Error extracting publication: {str(e)}"
                    logger.warning(error_msg)
                    errors.append(error_msg)
                    if progress_callback:
                        progress_callback('error', {'message': error_msg})

            next_url = None
            feed_links = feed_data.get('links', [])
            logger.debug(f"Total links in response: {len(feed_links)}")
            for link in feed_links:
                logger.debug(f"Link found: rel={link.get('rel') if isinstance(link, dict) else 'N/A'}, href={link.get('href') if isinstance(link, dict) else 'N/A'}")
                if isinstance(link, dict) and link.get('rel') == 'next':
                    next_url = link.get('href')
                    if next_url:
                        # Try urljoin first for relative URLs, but if it's already absolute it stays as-is
                        resolved_url = urljoin(url, next_url)
                        logger.info("Original next_url: %s", next_url)
                        logger.info("Resolved next_url: %s", resolved_url)
                        next_url = resolved_url
                        if "token=" in next_url:
                            current_auth = None
                    break

            if not next_url:
                logger.info("No next link found in page %s, pagination complete", pages_crawled)
                if progress_callback:
                    progress_callback('pagination_complete', {
                        'page_number': pages_crawled,
                        'message': 'No more pages to crawl'
                    })
            
            if next_url and (not max_pages or pages_crawled < max_pages):
                logger.info("Following pagination link to next page")
                if progress_callback:
                    progress_callback('pagination_found', {
                        'page_number': pages_crawled,
                        'next_url': next_url
                    })
                crawl_page(next_url)

        except requests.exceptions.HTTPError as e:
            if e.response.status_code == 401:
                error_msg = f"Authentication failed (401 Unauthorized) on page {pages_crawled + 1}. Verify your credentials are correct for this feed."
            elif e.response.status_code == 403:
                error_msg = f"Access forbidden (403 Forbidden) on page {pages_crawled + 1}. Your credentials may not have permission."
            else:
                error_msg = f"HTTP error on page {pages_crawled + 1}: {str(e)}"
            logger.error(error_msg)
            if progress_callback:
                progress_callback('error', {'message': error_msg})
            errors.append(error_msg)
            # Don't break on HTTP errors - continue with what we have
        except requests.RequestException as e:
            error_msg = f"Network error on {url}: {str(e)}"
            logger.error(error_msg)
            if progress_callback:
                progress_callback('error', {'message': error_msg})
            errors.append(error_msg)
            # Continue on network errors too
        except Exception as e:
            error_msg = f"Error processing {url}: {str(e)}"
            logger.error(error_msg)
            if progress_callback:
                progress_callback('error', {'message': error_msg})
            errors.append(error_msg)

    crawl_page(feed_url)

    stats = calculate_odl_inventory_stats(inventory)
    stats['pages_crawled'] = pages_crawled
    stats['total_publications'] = len(inventory)
    stats['errors_count'] = len(errors)

    return {
        'inventory': inventory,
        'stats': stats,
        'errors': errors
    }


def extract_odl_inventory_record(publication, base_url=None):
    metadata = publication.get('metadata', {})

    identifier = metadata.get('identifier', 'N/A')
    title = metadata.get('title', 'Untitled')
    author = extract_author(metadata)
    publisher = extract_publisher(metadata)

    formats = detect_odl_formats(publication)
    format_str = ', '.join(formats) if formats else 'Unknown'

    drm_schemes = detect_odl_drm_scheme(publication)
    drm_str = ', '.join(drm_schemes) if drm_schemes else 'Unknown'

    order_name, order_id = extract_order_info(publication)
    price = extract_price(publication)

    terms = extract_license_terms(publication)
    concurrency = terms.get('concurrency')
    copy_allowed = _format_bool(terms.get('copy_allowed'))
    print_allowed = _format_bool(terms.get('print_allowed'))

    return {
        'identifier': identifier,
        'title': title,
        'author': author,
        'publisher': publisher,
        'format': format_str,
        'drm': drm_str,
        'order_name': order_name,
        'order_id': order_id,
        'price': price,
        'concurrency': concurrency if concurrency is not None else 'Unknown',
        'copy_allowed': copy_allowed,
        'print_allowed': print_allowed
    }


def extract_order_info(publication):
    licenses = publication.get('licenses', [])
    if not isinstance(licenses, list) or not licenses:
        return 'Unknown', 'Unknown'

    license_obj = licenses[0]
    metadata = license_obj.get('metadata', {})

    order = metadata.get('order') or license_obj.get('order') or {}

    order_name = (
        order.get('name')
        or metadata.get('order_name')
        or metadata.get('orderName')
        or license_obj.get('order_name')
        or 'Unknown'
    )
    order_id = (
        order.get('id')
        or order.get('identifier')
        or metadata.get('order_id')
        or metadata.get('orderId')
        or license_obj.get('order_id')
        or 'Unknown'
    )

    return order_name, order_id


def extract_price(publication):
    licenses = publication.get('licenses', [])
    if not isinstance(licenses, list) or not licenses:
        return 'Unknown'

    license_obj = licenses[0]
    metadata = license_obj.get('metadata', {})
    price = metadata.get('price') or license_obj.get('price')

    return _format_price(price)


def _format_price(price):
    if price is None:
        return 'Unknown'

    if isinstance(price, dict):
        value = price.get('value') or price.get('amount') or price.get('price')
        currency = price.get('currency') or price.get('currencyCode')
        if value is not None and currency:
            return f"{value} {currency}"
        if value is not None:
            return str(value)
        return str(price)

    if isinstance(price, list):
        parts = [_format_price(item) for item in price if item is not None]
        return ', '.join([p for p in parts if p]) or 'Unknown'

    return str(price)


def _format_bool(value):
    if value is True:
        return 'Yes'
    if value is False:
        return 'No'
    return 'Unknown'


def calculate_odl_inventory_stats(inventory):
    format_counts = {}
    drm_counts = {}

    for record in inventory:
        formats = record.get('format', 'Unknown')
        format_counts[formats] = format_counts.get(formats, 0) + 1

        drm = record.get('drm', 'Unknown')
        drm_counts[drm] = drm_counts.get(drm, 0) + 1

    format_stats = sorted(
        [{'type': k, 'count': v} for k, v in format_counts.items()],
        key=lambda x: x['count'],
        reverse=True
    )

    drm_stats = sorted(
        [{'type': k, 'count': v} for k, v in drm_counts.items()],
        key=lambda x: x['count'],
        reverse=True
    )

    return {
        'format_counts': format_stats,
        'drm_counts': drm_stats,
        'unique_formats': len(format_counts),
        'unique_drm_types': len(drm_counts)
    }


def generate_odl_inventory_csv(inventory_data):
    output = io.StringIO()
    fieldnames = [
        'identifier',
        'title',
        'author',
        'publisher',
        'format',
        'drm',
        'order_name',
        'order_id',
        'price',
        'concurrency',
        'copy_allowed',
        'print_allowed'
    ]

    writer = csv.DictWriter(output, fieldnames=fieldnames, quoting=csv.QUOTE_MINIMAL)
    writer.writeheader()

    for record in inventory_data:
        row = {key: record.get(key, '') for key in fieldnames}
        writer.writerow(row)

    return output.getvalue()


def generate_odl_inventory_xml(inventory_data):
    root = Element('odl_inventory')
    root.set('total', str(len(inventory_data)))

    for record in inventory_data:
        pub_elem = SubElement(root, 'publication')
        for field in [
            'identifier', 'title', 'author', 'publisher', 'format', 'drm',
            'order_name', 'order_id', 'price', 'concurrency', 'copy_allowed', 'print_allowed'
        ]:
            child = SubElement(pub_elem, field)
            child.text = str(record.get(field, ''))

    xml_string = tostring(root, encoding='unicode')
    dom = minidom.parseString(xml_string)
    return dom.toprettyxml(indent="  ")


def generate_odl_inventory_excel(inventory_data):
    wb = Workbook()
    ws = wb.active
    ws.title = "ODL Inventory"

    header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    headers = [
        'identifier', 'title', 'author', 'publisher', 'format', 'drm',
        'order_name', 'order_id', 'price', 'concurrency', 'copy_allowed', 'print_allowed'
    ]

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header.replace('_', ' ').title())
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    for row_idx, record in enumerate(inventory_data, start=2):
        for col_idx, field in enumerate(headers, start=1):
            value = record.get(field, '')
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = cell_alignment

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 25
    ws.column_dimensions['H'].width = 25
    ws.column_dimensions['I'].width = 15
    ws.column_dimensions['J'].width = 15
    ws.column_dimensions['K'].width = 15
    ws.column_dimensions['L'].width = 15

    ws.freeze_panes = "A2"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return output.getvalue()
