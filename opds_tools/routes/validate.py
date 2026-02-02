from flask import Blueprint, request, render_template, flash, Response, redirect, url_for, stream_with_context
from opds_tools.util.palace_validator import validate_feed_url
import json
import io
import queue
import threading

from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

validate_bp = Blueprint("validate", __name__)

# Simple in-memory cache for last validation results
_last_validation = {
    "results": None,
    "feed_url": None,
    "max_pages": None,
}



@validate_bp.route("/validate-feed", methods=["GET", "POST"])
def validate_feed_view():
    results = {}
    feed_url = ""
    max_pages = None

    if request.method == "POST":
        action = request.form.get("action")

        if action == "clear":
            _last_validation["results"] = None
            _last_validation["feed_url"] = None
            _last_validation["max_pages"] = None
            return render_template("validate_feed.html", results={}, feed_url="", max_pages=None)

        feed_url = request.form.get("feed_url")
        max_pages_input = request.form.get("max_pages")

        if max_pages_input:
            try:
                max_pages = int(max_pages_input)
                if max_pages < 1:
                    max_pages = None
            except ValueError:
                max_pages = None

        download_json = request.form.get("download_json")

        if download_json and _last_validation["results"]:
            return Response(
                json.dumps(_last_validation["results"], indent=2),
                mimetype="application/json",
                headers={"Content-Disposition": "attachment;filename=validation.json"}
            )

        if feed_url:
            results = validate_feed_url(feed_url, max_pages=max_pages)

            _last_validation["results"] = results
            _last_validation["feed_url"] = feed_url
            _last_validation["max_pages"] = max_pages

    else:
        # GET request - show cached results if available
        if _last_validation["results"]:
            results = _last_validation["results"]
            feed_url = _last_validation["feed_url"]
            max_pages = _last_validation["max_pages"]

    return render_template("validate_feed.html", results=results, feed_url=feed_url, max_pages=max_pages)


@validate_bp.route("/validate-feed/pdf", methods=["GET"])
def validate_feed_pdf():
    """
    Generate a PDF report from the latest validation results.
    """
    if not _last_validation.get("results"):
        flash("No validation results available. Run a validation first.", "warning")
        return redirect(url_for("validate.validate_feed_view"))

    results = _last_validation["results"]
    feed_url = _last_validation.get("feed_url")
    max_pages = _last_validation.get("max_pages")

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph("OPDS Feed Validation Report", styles["Title"]))
    elements.append(Spacer(1, 12))

    elements.append(Paragraph(f"<b>Feed:</b> {feed_url or 'N/A'}", styles["Normal"]))
    elements.append(Paragraph(f"<b>Max Pages:</b> {max_pages or 'All'}", styles["Normal"]))
    elements.append(Spacer(1, 12))

    summary = results.get("summary", {})
    summary_data = [
        ["Pages Validated", summary.get("pages_validated", 0)],
        ["Publications Checked", summary.get("publication_count", 0)],
        ["Total Errors", summary.get("error_count", 0)],
        ["Total Warnings", summary.get("warning_count", 0)],
    ]
    summary_table = Table(summary_data, hAlign="LEFT")
    summary_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
    ]))
    elements.append(Paragraph("Summary", styles["Heading2"]))
    elements.append(summary_table)
    elements.append(Spacer(1, 12))

    # Build error summary by page
    feed_errors = results.get("feed_errors", [])
    pub_errors = results.get("publication_errors", [])
    feed_warnings = results.get("feed_warnings", [])
    pub_warnings = results.get("publication_warnings", [])
    
    if feed_errors or pub_errors or feed_warnings or pub_warnings:
        elements.append(Paragraph("Validation Issues Summary by Page", styles["Heading2"]))
        
        # Group errors and warnings by page
        error_by_page = {}
        for err in feed_errors:
            page = err.get("page_number", "Unknown")
            if page not in error_by_page:
                error_by_page[page] = {"schema": 0, "structure": 0, "publications": 0, "warnings": 0, "pub_names": []}
            if "Schema validation" in err.get("error", ""):
                error_by_page[page]["schema"] += 1
            else:
                error_by_page[page]["structure"] += 1
        
        for warn in feed_warnings:
            page = warn.get("page_number", "Unknown")
            if page not in error_by_page:
                error_by_page[page] = {"schema": 0, "structure": 0, "publications": 0, "warnings": 0, "pub_names": []}
            error_by_page[page]["warnings"] += 1
        
        for err in pub_errors:
            page = err.get("page_number", "Unknown")
            if page not in error_by_page:
                error_by_page[page] = {"schema": 0, "structure": 0, "publications": 0, "warnings": 0, "pub_names": []}
            error_by_page[page]["publications"] += 1
            error_by_page[page]["pub_names"].append(err.get("title", "Untitled"))
        
        for warn in pub_warnings:
            page = warn.get("page_number", "Unknown")
            if page not in error_by_page:
                error_by_page[page] = {"schema": 0, "structure": 0, "publications": 0, "warnings": 0, "pub_names": []}
            error_by_page[page]["warnings"] += 1
        
        # Create summary table
        page_summary_rows = [["Page", "Schema Errors", "Structure Errors", "Publication Errors", "Warnings", "Total"]]
        for page in sorted(error_by_page.keys()):
            data = error_by_page[page]
            total = data["schema"] + data["structure"] + data["publications"] + data["warnings"]
            page_summary_rows.append([
                f"Page {page}",
                str(data["schema"]),
                str(data["structure"]),
                str(data["publications"]),
                str(data["warnings"]),
                str(total)
            ])
        
        page_summary_table = Table(page_summary_rows, hAlign="LEFT", colWidths=[60, 75, 75, 90, 70, 50])
        page_summary_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.lightblue),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("ALIGN", (1, 0), (-1, -1), "CENTER"),
        ]))
        elements.append(page_summary_table)
        elements.append(Spacer(1, 12))

    feed_errors = results.get("feed_errors", [])
    if feed_errors:
        elements.append(Paragraph("Feed Page Errors", styles["Heading2"]))
        rows = [["Page", "URL", "Error"]]
        for err in feed_errors[:50]:
            rows.append([
                f"Page {err.get('page_number', '?')}",
                err.get("url", "")[:40] + "...",
                err.get("error", "")[:60] + "..."
            ])
        table = Table(rows, hAlign="LEFT", colWidths=[60, 200, 280])
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        elements.append(table)
        if len(feed_errors) > 50:
            elements.append(Paragraph("Showing first 50 feed errors.", styles["Italic"]))
        elements.append(Spacer(1, 12))

    pub_errors = results.get("publication_errors", [])
    if pub_errors:
        elements.append(Paragraph("Publication Errors", styles["Heading2"]))
        rows = [["Page", "Title", "Identifier", "Error"]]
        for pub in pub_errors[:50]:
            rows.append([
                f"Page {pub.get('page_number', '?')}",
                pub.get("title") or "Untitled",
                pub.get("identifier") or "",
                pub.get("error") or "",
            ])
        table = Table(rows, hAlign="LEFT", colWidths=[60, 160, 140, 180])
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        elements.append(table)
        if len(pub_errors) > 50:
            elements.append(Paragraph("Showing first 50 publication errors.", styles["Italic"]))
        elements.append(Spacer(1, 12))

    feed_warnings = results.get("feed_warnings", [])
    if feed_warnings:
        elements.append(Paragraph("Feed Page Warnings", styles["Heading2"]))
        rows = [["Page", "URL", "Warning", "Severity"]]
        for warn in feed_warnings[:50]:
            rows.append([
                f"Page {warn.get('page_number', '?')}",
                warn.get("url", "")[:35] + "...",
                warn.get("warning", "")[:55] + "...",
                warn.get("severity", "warning").upper()
            ])
        table = Table(rows, hAlign="LEFT", colWidths=[60, 180, 250, 50])
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.lightyellow),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        elements.append(table)
        if len(feed_warnings) > 50:
            elements.append(Paragraph("Showing first 50 feed warnings.", styles["Italic"]))
        elements.append(Spacer(1, 12))

    pub_warnings = results.get("publication_warnings", [])
    if pub_warnings:
        elements.append(Paragraph("Publication Warnings", styles["Heading2"]))
        rows = [["Page", "Title", "Warning", "Severity"]]
        for warn in pub_warnings[:50]:
            rows.append([
                f"Page {warn.get('page_number', '?')}",
                warn.get("title") or "Untitled",
                warn.get("warning") or "",
                warn.get("severity", "warning").upper()
            ])
        table = Table(rows, hAlign="LEFT", colWidths=[60, 200, 230, 50])
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.lightyellow),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        elements.append(table)
        if len(pub_warnings) > 50:
            elements.append(Paragraph("Showing first 50 publication warnings.", styles["Italic"]))

    doc.build(elements)
    pdf = buffer.getvalue()
    buffer.close()

    return Response(
        pdf,
        mimetype="application/pdf",
        headers={"Content-Disposition": "attachment;filename=opds_validation_report.pdf"},
    )


@validate_bp.route("/validate-feed/excel", methods=["GET"])
def validate_feed_excel():
    """
    Generate an Excel report from the latest validation results.
    """
    if not _last_validation.get("results"):
        flash("No validation results available. Run a validation first.", "warning")
        return redirect(url_for("validate.validate_feed_view"))

    results = _last_validation["results"]
    feed_url = _last_validation.get("feed_url")
    max_pages = _last_validation.get("max_pages")

    # Create workbook
    wb = Workbook()
    
    # ===== SUMMARY SHEET =====
    ws_summary = wb.active
    ws_summary.title = "Summary"
    
    # Define styles
    header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    info_fill = PatternFill(start_color="E8F4F8", end_color="E8F4F8", fill_type="solid")
    info_font = Font(bold=True, size=11)
    
    error_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
    warning_fill = PatternFill(start_color="FFF8DC", end_color="FFF8DC", fill_type="solid")
    
    cell_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    
    # Title
    ws_summary['A1'] = "OPDS Feed Validation Report"
    ws_summary['A1'].font = Font(bold=True, size=14)
    ws_summary.merge_cells('A1:B1')
    
    # Feed info
    row = 3
    ws_summary[f'A{row}'] = "Feed URL:"
    ws_summary[f'B{row}'] = feed_url or 'N/A'
    
    row += 1
    ws_summary[f'A{row}'] = "Max Pages:"
    ws_summary[f'B{row}'] = max_pages or 'All'
    
    # Summary statistics
    row += 2
    ws_summary[f'A{row}'] = "Metric"
    ws_summary[f'B{row}'] = "Value"
    ws_summary[f'A{row}'].fill = header_fill
    ws_summary[f'B{row}'].fill = header_fill
    ws_summary[f'A{row}'].font = header_font
    ws_summary[f'B{row}'].font = header_font
    
    summary = results.get("summary", {})
    metrics = [
        ("Pages Validated", summary.get("pages_validated", 0)),
        ("Publications Checked", summary.get("publication_count", 0)),
        ("Total Errors", summary.get("error_count", 0)),
        ("Total Warnings", summary.get("warning_count", 0)),
    ]
    
    for metric_name, metric_value in metrics:
        row += 1
        ws_summary[f'A{row}'] = metric_name
        ws_summary[f'B{row}'] = metric_value
        if "Error" in metric_name:
            ws_summary[f'A{row}'].fill = error_fill
            ws_summary[f'B{row}'].fill = error_fill
        elif "Warning" in metric_name:
            ws_summary[f'A{row}'].fill = warning_fill
            ws_summary[f'B{row}'].fill = warning_fill
        else:
            ws_summary[f'A{row}'].fill = info_fill
            ws_summary[f'B{row}'].fill = info_fill
    
    ws_summary.column_dimensions['A'].width = 25
    ws_summary.column_dimensions['B'].width = 50
    
    # ===== ISSUES BY PAGE SHEET =====
    ws_issues = wb.create_sheet("Issues by Page")
    
    # Build summary by page
    error_by_page = {}
    for err in results.get("feed_errors", []):
        page = err.get("page_number", "Unknown")
        if page not in error_by_page:
            error_by_page[page] = {"schema": 0, "structure": 0, "publications": 0, "warnings": 0}
        if "Schema validation" in err.get("error", ""):
            error_by_page[page]["schema"] += 1
        else:
            error_by_page[page]["structure"] += 1
    
    for warn in results.get("feed_warnings", []):
        page = warn.get("page_number", "Unknown")
        if page not in error_by_page:
            error_by_page[page] = {"schema": 0, "structure": 0, "publications": 0, "warnings": 0}
        error_by_page[page]["warnings"] += 1
    
    for err in results.get("publication_errors", []):
        page = err.get("page_number", "Unknown")
        if page not in error_by_page:
            error_by_page[page] = {"schema": 0, "structure": 0, "publications": 0, "warnings": 0}
        error_by_page[page]["publications"] += 1
    
    for warn in results.get("publication_warnings", []):
        page = warn.get("page_number", "Unknown")
        if page not in error_by_page:
            error_by_page[page] = {"schema": 0, "structure": 0, "publications": 0, "warnings": 0}
        error_by_page[page]["warnings"] += 1
    
    # Write headers
    headers = ["Page", "Schema Errors", "Structure Errors", "Publication Errors", "Warnings", "Total"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws_issues.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Write data
    for row_idx, page in enumerate(sorted(error_by_page.keys()), start=2):
        data = error_by_page[page]
        total = data["schema"] + data["structure"] + data["publications"] + data["warnings"]
        
        values = [
            f"Page {page}",
            data["schema"],
            data["structure"],
            data["publications"],
            data["warnings"],
            total
        ]
        
        for col_idx, value in enumerate(values, start=1):
            cell = ws_issues.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal="center", vertical="center")
    
    ws_issues.column_dimensions['A'].width = 12
    ws_issues.column_dimensions['B'].width = 15
    ws_issues.column_dimensions['C'].width = 18
    ws_issues.column_dimensions['D'].width = 20
    ws_issues.column_dimensions['E'].width = 12
    ws_issues.column_dimensions['F'].width = 10
    
    # ===== FEED ERRORS SHEET =====
    if results.get("feed_errors"):
        ws_errors = wb.create_sheet("Feed Errors")
        headers = ["Page", "URL", "Error", "Details"]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws_errors.cell(row=1, column=col_idx, value=header)
            cell.fill = error_fill
            cell.font = Font(bold=True, color="8B0000", size=11)
            cell.alignment = header_alignment
        
        for row_idx, err in enumerate(results.get("feed_errors", [])[:100], start=2):
            ws_errors.cell(row=row_idx, column=1, value=f"Page {err.get('page_number', '?')}")
            ws_errors.cell(row=row_idx, column=2, value=err.get("url", ""))
            ws_errors.cell(row=row_idx, column=3, value=err.get("error", ""))
            details = json.dumps(err.get("details", []), indent=2) if err.get("details") else ""
            ws_errors.cell(row=row_idx, column=4, value=details)
        
        ws_errors.column_dimensions['A'].width = 12
        ws_errors.column_dimensions['B'].width = 40
        ws_errors.column_dimensions['C'].width = 35
        ws_errors.column_dimensions['D'].width = 30
    
    # ===== PUBLICATION ERRORS SHEET =====
    if results.get("publication_errors"):
        ws_pub_errors = wb.create_sheet("Publication Errors")
        headers = ["Page", "Title", "Identifier", "Author", "Error"]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws_pub_errors.cell(row=1, column=col_idx, value=header)
            cell.fill = error_fill
            cell.font = Font(bold=True, color="8B0000", size=11)
            cell.alignment = header_alignment
        
        for row_idx, pub in enumerate(results.get("publication_errors", [])[:100], start=2):
            ws_pub_errors.cell(row=row_idx, column=1, value=f"Page {pub.get('page_number', '?')}")
            ws_pub_errors.cell(row=row_idx, column=2, value=pub.get("title", "Untitled"))
            ws_pub_errors.cell(row=row_idx, column=3, value=pub.get("identifier", ""))
            ws_pub_errors.cell(row=row_idx, column=4, value=pub.get("author", ""))
            ws_pub_errors.cell(row=row_idx, column=5, value=pub.get("error", ""))
        
        ws_pub_errors.column_dimensions['A'].width = 12
        ws_pub_errors.column_dimensions['B'].width = 35
        ws_pub_errors.column_dimensions['C'].width = 30
        ws_pub_errors.column_dimensions['D'].width = 25
        ws_pub_errors.column_dimensions['E'].width = 40
    
    # ===== FEED WARNINGS SHEET =====
    if results.get("feed_warnings"):
        ws_feed_warn = wb.create_sheet("Feed Warnings")
        headers = ["Page", "URL", "Warning", "Severity"]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws_feed_warn.cell(row=1, column=col_idx, value=header)
            cell.fill = warning_fill
            cell.font = Font(bold=True, color="B8860B", size=11)
            cell.alignment = header_alignment
        
        for row_idx, warn in enumerate(results.get("feed_warnings", [])[:100], start=2):
            ws_feed_warn.cell(row=row_idx, column=1, value=f"Page {warn.get('page_number', '?')}")
            ws_feed_warn.cell(row=row_idx, column=2, value=warn.get("url", ""))
            ws_feed_warn.cell(row=row_idx, column=3, value=warn.get("warning", ""))
            ws_feed_warn.cell(row=row_idx, column=4, value=warn.get("severity", "warning").upper())
        
        ws_feed_warn.column_dimensions['A'].width = 12
        ws_feed_warn.column_dimensions['B'].width = 40
        ws_feed_warn.column_dimensions['C'].width = 40
        ws_feed_warn.column_dimensions['D'].width = 15
    
    # ===== PUBLICATION WARNINGS SHEET =====
    if results.get("publication_warnings"):
        ws_pub_warn = wb.create_sheet("Publication Warnings")
        headers = ["Page", "Title", "Identifier", "Warning", "Severity"]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws_pub_warn.cell(row=1, column=col_idx, value=header)
            cell.fill = warning_fill
            cell.font = Font(bold=True, color="B8860B", size=11)
            cell.alignment = header_alignment
        
        for row_idx, warn in enumerate(results.get("publication_warnings", [])[:100], start=2):
            ws_pub_warn.cell(row=row_idx, column=1, value=f"Page {warn.get('page_number', '?')}")
            ws_pub_warn.cell(row=row_idx, column=2, value=warn.get("title", "Untitled"))
            ws_pub_warn.cell(row=row_idx, column=3, value=warn.get("identifier", ""))
            ws_pub_warn.cell(row=row_idx, column=4, value=warn.get("warning", ""))
            ws_pub_warn.cell(row=row_idx, column=5, value=warn.get("severity", "warning").upper())
        
        ws_pub_warn.column_dimensions['A'].width = 12
        ws_pub_warn.column_dimensions['B'].width = 35
        ws_pub_warn.column_dimensions['C'].width = 30
        ws_pub_warn.column_dimensions['D'].width = 40
        ws_pub_warn.column_dimensions['E'].width = 15
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return Response(
        output.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment;filename=opds_validation_report.xlsx"},
    )


@validate_bp.route("/validate-feed/stream", methods=["GET"])
def validate_feed_stream():
    """
    Stream progress updates for feed validation using Server-Sent Events.
    """
    feed_url = request.args.get("feed_url", "").strip()
    max_pages_input = request.args.get("max_pages", "").strip()
    
    max_pages = None
    if max_pages_input:
        try:
            max_pages = int(max_pages_input)
            if max_pages < 1:
                max_pages = None
        except ValueError:
            max_pages = None
    
    def generate():
        """Generator for Server-Sent Events."""
        global _last_validation
        
        # Create a queue for progress events
        progress_queue = queue.Queue()
        
        def run_validation():
            """Run validation in background thread."""
            try:
                def on_progress(event_type, data):
                    progress_queue.put({'type': event_type, **data})
                
                results = validate_feed_url(
                    feed_url,
                    max_pages=max_pages,
                    progress_callback=on_progress
                )
                
                # Cache results
                _last_validation['results'] = results
                _last_validation['feed_url'] = feed_url
                _last_validation['max_pages'] = max_pages
                
                # Signal completion
                progress_queue.put({
                    'type': 'complete',
                    'summary': results['summary'],
                    'error_count': results['summary']['error_count'],
                    'feed_error_count': len(results.get('feed_errors', [])),
                    'publication_error_count': len(results.get('publication_errors', []))
                })
                
            except Exception as e:
                import traceback
                traceback.print_exc()
                progress_queue.put({'type': 'error', 'message': str(e)})
        
        # Start background thread
        thread = threading.Thread(target=run_validation)
        thread.daemon = True
        thread.start()
        
        # Send initial status
        yield f"data: {json.dumps({'type': 'started', 'feed_url': feed_url})}\n\n"
        
        # Stream events from queue
        while True:
            try:
                event = progress_queue.get(timeout=60)  # 60 second timeout
                yield f"data: {json.dumps(event)}\n\n"
                
                # Stop if complete or error
                if event['type'] in ['complete', 'error']:
                    break
                    
            except queue.Empty:
                # Send keepalive
                yield f": keepalive\n\n"
    
    return Response(
        stream_with_context(generate()),
        mimetype="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "X-Accel-Buffering": "no",
            "Connection": "keep-alive"
        }
    )
