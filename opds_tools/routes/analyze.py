from flask import Blueprint, request, render_template, Response, redirect, url_for, flash, stream_with_context
import json
import time
import io
import queue
import threading

from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import BarChart, PieChart, Reference

from opds_tools.util.feed_analyzer import analyze_feed_url

analyze_bp = Blueprint("analyze", __name__)

# Simple in-memory cache for the last analysis result
# In production, you'd use Redis or a database
_last_analysis = {
    'results': None,
    'feed_url': None,
    'max_pages': None,
    'in_progress': False,
    'started_at': None,
}


@analyze_bp.route("/analyze-feed", methods=["GET", "POST"])
def analyze_feed_view():
    """
    Analyze OPDS feed for format and DRM statistics.
    """
    global _last_analysis
    
    results = {}
    feed_url = ""
    max_pages = None

    if request.method == "POST":
        action = request.form.get("action")

        if action == "clear":
            _last_analysis = {'results': None, 'feed_url': None, 'max_pages': None, 'in_progress': False, 'started_at': None}
            return render_template("analyze_feed.html", results={}, feed_url="", max_pages=None)

        feed_url = request.form.get("feed_url")
        max_pages_input = request.form.get("max_pages")

        if max_pages_input:
            try:
                max_pages = int(max_pages_input)
                if max_pages < 1:
                    max_pages = None
            except ValueError:
                max_pages = None

        download_json = request.form.get("download_json")
        
        # Handle JSON download from cached results
        if download_json and _last_analysis['results']:
            print(f"📥 Serving JSON download from cached results")
            return Response(
                json.dumps(_last_analysis['results'], indent=2),
                mimetype="application/json",
                headers={"Content-Disposition": "attachment;filename=feed_analysis.json"}
            )
        
        if feed_url:
            # Run analysis
            try:
                print(f"\n🔬 Starting analysis for: {feed_url}")
                print(f"   Max pages: {max_pages or 'unlimited'}")
                print(f"   ⏱️  This may take several minutes for large feeds...")

                _last_analysis['in_progress'] = True
                _last_analysis['started_at'] = time.time()
                _last_analysis['feed_url'] = feed_url
                _last_analysis['max_pages'] = max_pages
                _last_analysis['results'] = None
                
                results = analyze_feed_url(feed_url, max_pages=max_pages)
                
                print(f"\n✅ Analysis complete!")
                print(f"   Total publications: {results.get('summary', {}).get('total_publications', 0)}")
                print(f"   Pages analyzed: {results.get('summary', {}).get('pages_analyzed', 0)}")
                print(f"   Format combinations found: {len(results.get('format_combo_stats', []))}")
                
                # Limit page_stats for template rendering to avoid massive HTML
                # Only show first 20 and last 5 pages in detail
                page_stats_count = len(results.get('page_stats', []))
                if page_stats_count > 25:
                    print(f"   ⚠️  Limiting page details display: {page_stats_count} pages total, showing first 20 + last 5")
                    page_stats = results.get('page_stats', [])
                    results['page_stats_display'] = page_stats[:20] + page_stats[-5:]
                    results['page_stats_truncated'] = True
                    results['page_stats_total'] = page_stats_count
                else:
                    results['page_stats_display'] = results.get('page_stats', [])
                    results['page_stats_truncated'] = False
                
                # Cache the results
                _last_analysis['results'] = results
                _last_analysis['feed_url'] = feed_url
                _last_analysis['max_pages'] = max_pages
                _last_analysis['in_progress'] = False
                _last_analysis['started_at'] = None
                
            except Exception as e:
                print(f"\n❌ Error during analysis: {str(e)}")
                import traceback
                traceback.print_exc()
                results = {
                    "error": str(e),
                    "summary": {
                        "total_publications": 0,
                        "pages_analyzed": 0,
                        "pages_with_errors": 0
                    }
                }
                results['page_stats_display'] = []
                results['page_stats_truncated'] = False
                _last_analysis['in_progress'] = False
                _last_analysis['started_at'] = None

    else:
        # GET request - show cached results if available
        if _last_analysis['results']:
            print(f"📋 Displaying cached analysis results")
            results = _last_analysis['results']
            feed_url = _last_analysis['feed_url']
            max_pages = _last_analysis['max_pages']
        elif _last_analysis['in_progress']:
            print(f"⏳ Analysis still in progress...")
            feed_url = _last_analysis['feed_url'] or ""
            max_pages = _last_analysis['max_pages']
            results = {"in_progress": True, "summary": {}}

    print(f"🎨 Rendering template with results...")
    response = render_template("analyze_feed.html", results=results, feed_url=feed_url, max_pages=max_pages)
    print(f"✅ Template rendered successfully, sending response to browser")
    return response


@analyze_bp.route("/analyze-feed/pdf", methods=["GET"])
def analyze_feed_pdf():
    """
    Generate a PDF report from the latest analysis results.
    """
    global _last_analysis

    if not _last_analysis.get("results"):
        flash("No analysis results available. Run an analysis first.", "warning")
        return redirect(url_for("analyze.analyze_feed_view"))

    results = _last_analysis["results"]
    feed_url = _last_analysis.get("feed_url")
    max_pages = _last_analysis.get("max_pages")

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph("OPDS Feed Analysis Report", styles["Title"]))
    elements.append(Spacer(1, 12))

    elements.append(Paragraph(f"<b>Feed:</b> {feed_url or 'N/A'}", styles["Normal"]))
    elements.append(Paragraph(f"<b>Max Pages:</b> {max_pages or 'All'}", styles["Normal"]))
    elements.append(Spacer(1, 12))

    summary = results.get("summary", {})
    type_counts = summary.get("publication_type_counts", {})
    type_percentages = summary.get("publication_type_percentages", {})
    summary_data = [
        ["Total Publications", summary.get("total_publications", 0)],
        ["Pages Analyzed", summary.get("pages_analyzed", 0)],
        ["Pages With Errors", summary.get("pages_with_errors", 0)],
        ["Unique Formats", summary.get("unique_formats", 0)],
        ["Unique Format Combinations", summary.get("unique_format_combinations", 0)],
        ["Unique DRM Types", summary.get("unique_drm_types", 0)],
        ["Bearer Token DRM Publications", summary.get("bearer_token_publications", 0)],
        ["Audiobook Publications", summary.get("audiobook_publications", 0)],
        ["Publications with Samples", summary.get("sample_publications", 0)],
        ["Publication Types - Book", f"{type_counts.get('Book', 0)} ({type_percentages.get('Book', 0)}%)"],
        ["Publication Types - Audiobook", f"{type_counts.get('Audiobook', 0)} ({type_percentages.get('Audiobook', 0)}%)"],
        ["Publication Types - Periodical", f"{type_counts.get('Periodical', 0)} ({type_percentages.get('Periodical', 0)}%)"],
        ["Publication Types - Other", f"{type_counts.get('Other', 0)} ({type_percentages.get('Other', 0)}%)"],
    ]
    summary_table = Table(summary_data, hAlign="LEFT")
    summary_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
    ]))
    elements.append(Paragraph("Summary", styles["Heading2"]))
    elements.append(summary_table)
    elements.append(Spacer(1, 12))

    elements.append(Paragraph("Format Availability (Combinations)", styles["Heading2"]))
    combo_rows = [["Combination", "Count", "% of Collection"]]
    for item in results.get("format_combo_stats", []):
        combo_rows.append([item.get("combination"), item.get("count"), f"{item.get('percentage', 0)}%"])
    combo_table = Table(combo_rows, hAlign="LEFT")
    combo_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
    ]))
    elements.append(combo_table)
    elements.append(Spacer(1, 12))

    elements.append(Paragraph("Format Distribution", styles["Heading2"]))
    format_rows = [["Format", "Count", "% of Collection"]]
    total_pubs = summary.get("total_publications", 0) or 1
    for fmt, count in results.get("format_counts", {}).items():
        pct = (count / total_pubs) * 100
        format_rows.append([fmt, count, f"{pct:.1f}%"])
    format_table = Table(format_rows, hAlign="LEFT")
    format_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
    ]))
    elements.append(format_table)
    elements.append(Spacer(1, 12))

    drm_counts = results.get("drm_counts", {})
    if drm_counts:
        elements.append(Paragraph("DRM Distribution (EPUB Only)", styles["Heading2"]))
        drm_rows = [["DRM Type", "Count", "% of EPUBs", "% of Collection"]]
        epub_total = results.get("format_counts", {}).get("EPUB", 0) or 1
        for drm, count in drm_counts.items():
            if drm == "N/A":
                continue
            pct_epub = (count / epub_total) * 100
            pct_all = (count / total_pubs) * 100
            drm_rows.append([drm, count, f"{pct_epub:.1f}%", f"{pct_all:.1f}%"])
        drm_table = Table(drm_rows, hAlign="LEFT")
        drm_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        elements.append(drm_table)
        elements.append(Spacer(1, 12))

    elements.append(Paragraph("Format + DRM Distribution", styles["Heading2"]))
    combined_rows = [["Format", "DRM", "Count", "% of Collection"]]
    for item in results.get("combined_stats", []):
        combined_rows.append([
            item.get("format"),
            item.get("drm"),
            item.get("count"),
            f"{item.get('percentage', 0)}%",
        ])
    combined_table = Table(combined_rows, hAlign="LEFT")
    combined_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
    ]))
    elements.append(combined_table)

    doc.build(elements)
    pdf = buffer.getvalue()
    buffer.close()

    return Response(
        pdf,
        mimetype="application/pdf",
        headers={"Content-Disposition": "attachment;filename=opds_analysis_report.pdf"},
    )


@analyze_bp.route("/analyze-feed/excel", methods=["GET"])
def analyze_feed_excel():
    """
    Generate an Excel report from the latest analysis results.
    """
    global _last_analysis

    if not _last_analysis.get("results"):
        flash("No analysis results available. Run an analysis first.", "warning")
        return redirect(url_for("analyze.analyze_feed_view"))

    results = _last_analysis["results"]
    feed_url = _last_analysis.get("feed_url")
    max_pages = _last_analysis.get("max_pages")

    # Create workbook
    wb = Workbook()
    
    # ===== SUMMARY SHEET =====
    ws_summary = wb.active
    ws_summary.title = "Summary"
    
    # Define styles
    header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    info_fill = PatternFill(start_color="E8F4F8", end_color="E8F4F8", fill_type="solid")
    info_font = Font(bold=True, size=11)
    
    accent_fill = PatternFill(start_color="FFF8DC", end_color="FFF8DC", fill_type="solid")
    
    cell_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    
    # Title
    ws_summary['A1'] = "OPDS Feed Analysis Report"
    ws_summary['A1'].font = Font(bold=True, size=14)
    ws_summary.merge_cells('A1:B1')
    
    # Feed info
    row = 3
    ws_summary[f'A{row}'] = "Feed URL:"
    ws_summary[f'B{row}'] = feed_url or 'N/A'
    
    row += 1
    ws_summary[f'A{row}'] = "Max Pages:"
    ws_summary[f'B{row}'] = max_pages or 'All'
    
    # Summary statistics
    row += 2
    ws_summary[f'A{row}'] = "Metric"
    ws_summary[f'B{row}'] = "Value"
    ws_summary[f'A{row}'].fill = header_fill
    ws_summary[f'B{row}'].fill = header_fill
    ws_summary[f'A{row}'].font = header_font
    ws_summary[f'B{row}'].font = header_font
    
    summary = results.get("summary", {})
    type_counts = summary.get("publication_type_counts", {})
    type_percentages = summary.get("publication_type_percentages", {})
    
    metrics = [
        ("Total Publications", summary.get("total_publications", 0)),
        ("Pages Analyzed", summary.get("pages_analyzed", 0)),
        ("Pages With Errors", summary.get("pages_with_errors", 0)),
        ("Unique Formats", summary.get("unique_formats", 0)),
        ("Unique Format Combinations", summary.get("unique_format_combinations", 0)),
        ("Unique DRM Types", summary.get("unique_drm_types", 0)),
        ("Bearer Token DRM Publications", summary.get("bearer_token_publications", 0)),
        ("Audiobook Publications", summary.get("audiobook_publications", 0)),
        ("Publications with Samples", summary.get("sample_publications", 0)),
        ("", ""),  # Empty row
        ("Publication Types:", ""),
        ("  Book", f"{type_counts.get('Book', 0)} ({type_percentages.get('Book', 0)}%)"),
        ("  Audiobook", f"{type_counts.get('Audiobook', 0)} ({type_percentages.get('Audiobook', 0)}%)"),
        ("  Periodical", f"{type_counts.get('Periodical', 0)} ({type_percentages.get('Periodical', 0)}%)"),
        ("  Other", f"{type_counts.get('Other', 0)} ({type_percentages.get('Other', 0)}%)"),
    ]
    
    for metric_name, metric_value in metrics:
        row += 1
        ws_summary[f'A{row}'] = metric_name
        ws_summary[f'B{row}'] = metric_value
        if metric_name and not metric_name.startswith("  "):
            ws_summary[f'A{row}'].fill = info_fill
            ws_summary[f'B{row}'].fill = info_fill
    
    ws_summary.column_dimensions['A'].width = 35
    ws_summary.column_dimensions['B'].width = 50
    
    # ===== FORMAT COMBINATIONS SHEET =====
    ws_formats = wb.create_sheet("Format Combinations")
    
    headers = ["Format Combination", "Count", "% of Collection"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws_formats.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    for row_idx, item in enumerate(results.get("format_combo_stats", []), start=2):
        ws_formats.cell(row=row_idx, column=1, value=item.get("combination", ""))
        ws_formats.cell(row=row_idx, column=2, value=item.get("count", 0))
        ws_formats.cell(row=row_idx, column=3, value=f"{item.get('percentage', 0)}%")
        
        # Highlight multi-format rows
        if "+" in item.get("combination", ""):
            for col in range(1, 4):
                ws_formats.cell(row=row_idx, column=col).fill = accent_fill
    
    ws_formats.column_dimensions['A'].width = 40
    ws_formats.column_dimensions['B'].width = 15
    ws_formats.column_dimensions['C'].width = 20
    
    # ===== DRM DISTRIBUTION SHEET =====
    ws_drm = wb.create_sheet("DRM Distribution")
    
    # DRM counts for all formats
    headers = ["DRM Type", "Count", "% of Collection"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws_drm.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    drm_counts = results.get("drm_counts", {})
    total_pubs = summary.get("total_publications", 0) or 1
    row_idx = 2
    
    for drm, count in sorted(drm_counts.items(), key=lambda x: x[1], reverse=True):
        if drm != "N/A":
            ws_drm.cell(row=row_idx, column=1, value=drm)
            ws_drm.cell(row=row_idx, column=2, value=count)
            ws_drm.cell(row=row_idx, column=3, value=f"{(count / total_pubs * 100):.1f}%")
            row_idx += 1
    
    # Add spacing and EPUB-specific DRM stats
    row_idx += 2
    ws_drm.cell(row=row_idx, column=1, value="DRM Distribution (EPUB Only)")
    ws_drm.cell(row=row_idx, column=1).font = Font(bold=True, size=11)
    ws_drm.merge_cells(f'A{row_idx}:D{row_idx}')
    
    row_idx += 1
    headers_epub = ["DRM Type", "Count", "% of EPUBs", "% of Collection"]
    for col_idx, header in enumerate(headers_epub, start=1):
        cell = ws_drm.cell(row=row_idx, column=col_idx, value=header)
        cell.fill = info_fill
        cell.font = Font(bold=True, size=10)
        cell.alignment = header_alignment
    
    epub_total = results.get("format_counts", {}).get("EPUB", 0) or 1
    for drm, count in sorted(drm_counts.items(), key=lambda x: x[1], reverse=True):
        if drm != "N/A":
            row_idx += 1
            pct_epub = (count / epub_total) * 100
            pct_all = (count / total_pubs) * 100
            ws_drm.cell(row=row_idx, column=1, value=drm)
            ws_drm.cell(row=row_idx, column=2, value=count)
            ws_drm.cell(row=row_idx, column=3, value=f"{pct_epub:.1f}%")
            ws_drm.cell(row=row_idx, column=4, value=f"{pct_all:.1f}%")
    
    ws_drm.column_dimensions['A'].width = 25
    ws_drm.column_dimensions['B'].width = 15
    ws_drm.column_dimensions['C'].width = 18
    ws_drm.column_dimensions['D'].width = 20
    
    # ===== FORMAT + DRM COMBINATIONS SHEET =====
    ws_combined = wb.create_sheet("Format + DRM Details")
    
    headers = ["Format", "DRM", "Count", "% of Collection"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws_combined.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    for row_idx, item in enumerate(results.get("combined_stats", []), start=2):
        ws_combined.cell(row=row_idx, column=1, value=item.get("format", ""))
        ws_combined.cell(row=row_idx, column=2, value=item.get("drm", ""))
        ws_combined.cell(row=row_idx, column=3, value=item.get("count", 0))
        ws_combined.cell(row=row_idx, column=4, value=f"{item.get('percentage', 0)}%")
    
    ws_combined.column_dimensions['A'].width = 20
    ws_combined.column_dimensions['B'].width = 25
    ws_combined.column_dimensions['C'].width = 15
    ws_combined.column_dimensions['D'].width = 20
    
    # ===== PUBLICATION TYPES SHEET =====
    ws_types = wb.create_sheet("Publication Types")
    
    headers = ["Type", "Count", "Percentage"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws_types.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    type_data = [
        ("Book", type_counts.get('Book', 0), type_percentages.get('Book', 0)),
        ("Audiobook", type_counts.get('Audiobook', 0), type_percentages.get('Audiobook', 0)),
        ("Periodical", type_counts.get('Periodical', 0), type_percentages.get('Periodical', 0)),
        ("Other", type_counts.get('Other', 0), type_percentages.get('Other', 0)),
    ]
    
    for row_idx, (type_name, count, percentage) in enumerate(type_data, start=2):
        ws_types.cell(row=row_idx, column=1, value=type_name)
        ws_types.cell(row=row_idx, column=2, value=count)
        ws_types.cell(row=row_idx, column=3, value=f"{percentage}%")
    
    ws_types.column_dimensions['A'].width = 20
    ws_types.column_dimensions['B'].width = 15
    ws_types.column_dimensions['C'].width = 15
    
    # ===== PAGE STATISTICS SHEET (if available) =====
    if results.get("page_stats"):
        ws_pages = wb.create_sheet("Page Statistics")
        
        headers = ["Page #", "URL", "Publications", "Errors"]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws_pages.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        for page_idx, page_stat in enumerate(results.get("page_stats", [])[:200], start=1):
            row_idx = page_idx + 1  # +1 for header row
            ws_pages.cell(row=row_idx, column=1, value=page_idx)
            ws_pages.cell(row=row_idx, column=2, value=page_stat.get("url", ""))
            ws_pages.cell(row=row_idx, column=3, value=page_stat.get("publication_count", 0))
            error_msg = page_stat.get("error", "")
            ws_pages.cell(row=row_idx, column=4, value=error_msg if error_msg else "")
        
        ws_pages.column_dimensions['A'].width = 10
        ws_pages.column_dimensions['B'].width = 50
        ws_pages.column_dimensions['C'].width = 15
        ws_pages.column_dimensions['D'].width = 40
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return Response(
        output.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment;filename=opds_analysis_report.xlsx"},
    )


@analyze_bp.route("/analyze-feed/stream", methods=["GET"])
def analyze_feed_stream():
    """
    Stream progress updates for feed analysis using Server-Sent Events.
    """
    feed_url = request.args.get("feed_url", "").strip()
    max_pages_input = request.args.get("max_pages", "").strip()
    
    max_pages = None
    if max_pages_input:
        try:
            max_pages = int(max_pages_input)
            if max_pages < 1:
                max_pages = None
        except ValueError:
            max_pages = None
    
    def generate():
        """Generator for Server-Sent Events."""
        global _last_analysis
        
        _last_analysis['in_progress'] = True
        _last_analysis['started_at'] = time.time()
        
        # Create a queue for progress events
        progress_queue = queue.Queue()
        
        def run_analysis():
            """Run analysis in background thread."""
            try:
                def on_progress(event_type, data):
                    progress_queue.put({'type': event_type, **data})
                
                results = analyze_feed_url(
                    feed_url,
                    max_pages=max_pages,
                    progress_callback=on_progress
                )
                
                # Limit page_stats for template rendering
                page_stats_count = len(results.get('page_stats', []))
                if page_stats_count > 25:
                    page_stats = results.get('page_stats', [])
                    results['page_stats_display'] = page_stats[:20] + page_stats[-5:]
                    results['page_stats_truncated'] = True
                    results['page_stats_total'] = page_stats_count
                else:
                    results['page_stats_display'] = results.get('page_stats', [])
                    results['page_stats_truncated'] = False
                
                # Cache results
                _last_analysis['results'] = results
                _last_analysis['feed_url'] = feed_url
                _last_analysis['max_pages'] = max_pages
                _last_analysis['in_progress'] = False
                _last_analysis['started_at'] = None
                
                # Signal completion
                progress_queue.put({
                    'type': 'complete',
                    'summary': results.get('summary', {}),
                    'total_publications': results.get('summary', {}).get('total_publications', 0),
                    'pages_analyzed': results.get('summary', {}).get('pages_analyzed', 0)
                })
                
            except Exception as e:
                print(f"Error during analysis: {e}")
                import traceback
                traceback.print_exc()
                _last_analysis['in_progress'] = False
                _last_analysis['started_at'] = None
                progress_queue.put({'type': 'error', 'message': str(e)})
        
        # Start background thread
        thread = threading.Thread(target=run_analysis)
        thread.daemon = True
        thread.start()
        
        # Stream events from queue
        while True:
            try:
                event = progress_queue.get(timeout=60)  # 60 second timeout
                event_json = json.dumps(event)
                yield f"data: {event_json}\n\n"
                
                # Stop if complete or error
                if event['type'] in ['complete', 'error']:
                    break
                    
            except queue.Empty:
                # Send keepalive
                yield f": keepalive\n\n"
    
    return Response(
        stream_with_context(generate()),
        mimetype="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "X-Accel-Buffering": "no",
            "Connection": "keep-alive"
        }
    )
